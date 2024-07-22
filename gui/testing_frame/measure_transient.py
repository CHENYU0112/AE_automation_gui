import os
import time
from datetime import datetime
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment
from config import *
import pyvisa
from pverifyDrivers import _scope, powsup, daq, load

def transient(selected_ic, power_supply_settings, scope_settings, scope_persistence, scope_us_div, load_settings, protection, pass_fail_criteria, instrument_manager):
    print(f"Debug: transient function started for {selected_ic}")
    results = {}
    print("Transient test started")

    # Create result folder
    result_folder = os.path.join("results", "transient")
    os.makedirs(result_folder, exist_ok=True)

    # Generate unique filename for this test
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    test_name = f"Transient_Test_{selected_ic}_{timestamp}"
    excel_file = os.path.join(result_folder, f"{test_name}.xlsx")

    try:
        instruments = instrument_manager.initialize_instruments()
        
        electronic_load = instruments['load']
        power_supply = instruments['supply']
        sc = instruments['o_scope']
        
        power_supply_channel = power_supply.GetChannel(power_supply_settings['vin_channel'])

        # Set up power supply
        power_supply_channel.Configure_VoltageLevel(Level=power_supply_settings['vin'], CurrentLimit=power_supply_settings['iin'])
        power_supply_channel.Enable(True)

        # Find channel numbers for SW, Vout, and Iout
        sw_ch_num = next(ch for ch, signal in scope_settings.items() if signal == 'SW')
        vout_ch_num = next(ch for ch, signal in scope_settings.items() if signal == 'output V')
        iout_ch_num = next(ch for ch, signal in scope_settings.items() if signal == 'output I')

        # Setup oscilloscope channels
        sw_ch = sc.GetChannel(int(sw_ch_num[-1]))
        vout_ch = sc.GetChannel(int(vout_ch_num[-1]))
        iout_ch = sc.GetChannel(int(iout_ch_num[-1]))

        # Enable channels
        sw_ch.Enable(True)
        vout_ch.Enable(True)
        iout_ch.Enable(True)

        # Set up channels with provided settings and adjust vertical positions
        sw_ch.ProbeSetup(Coupling='DC', Bandwidth='20E+06', Vrange=5, Impedance=50, Position=3.5)
        iout_ch.ProbeSetup(Coupling='DC', Bandwidth='120E+06', Vrange=50, Impedance=1e+6, Position=0)
        vout_ch.ProbeSetup(Coupling='AC', Bandwidth='20E+06', Vrange=0.2, Impedance=1e+6, Position=-3.5)

        # Set channel labels
        sc.set_channel_label(ChannelIndex=int(sw_ch_num[-1]), Label="SW", Xpos="5", Ypos="80")
        sc.set_channel_label(ChannelIndex=int(iout_ch_num[-1]), Label="Iout", Xpos="5", Ypos="50")
        sc.set_channel_label(ChannelIndex=int(vout_ch_num[-1]), Label="Vout", Xpos="5", Ypos="20")

        # Set horizontal scale
        print(f"Debug: About to set horizontal scale to {scope_us_div}us")
        sc._write(f'HOR:SCALE {scope_us_div}e-6')  
        
        # Set trigger on the current channel for load transition
        trigger_level = (load_settings['i_high'] + load_settings['i_low']) / 2
        sc.Trigger_Edge(Level=trigger_level, Slope='RISE', Coupling='DC', ChannelIndex=int(iout_ch_num[-1]))

        sc.set_acquisition_mode('HIRes')
        sc.set_record_length(10000000)  # 10M points for higher resolution

        # Set measurements
        sc.set_measurement(MeasureIndex=1, MeasureType='MAXIMUM', State='ON', Source=int(vout_ch_num[-1]))
        sc.set_measurement(MeasureIndex=2, MeasureType='MINIMUM', State='ON', Source=int(vout_ch_num[-1]))
        sc.set_measurement(MeasureIndex=3, MeasureType='RISETIME', State='ON', Source=int(iout_ch_num[-1]))
        sc.set_measurement(MeasureIndex=4, MeasureType='FALLTIME', State='ON', Source=int(iout_ch_num[-1]))

        
        
        # Enable display overlay
        sc.display_overlay()
        
        # Setup electronic load for dynamic mode
        electronic_load.set_mode('CCDH')
        electronic_load._write(f'CURR:DYN:L1 {load_settings["i_low"]}')
        electronic_load._write(f'CURR:DYN:L2 {load_settings["i_high"]}')
        electronic_load._write(f'CURR:DYN:RISE MAX')
        electronic_load._write(f'CURR:DYN:FALL MAX')
        electronic_load._write(f'CURR:DYN:T1 {load_settings["low_time"]}us')
        electronic_load._write(f'CURR:DYN:T2 {load_settings["high_time"]}us')
        electronic_load._write('CURR:DYN:REP 0')

        # Start the load cycling
        electronic_load.output('ON')

        # Allow some time for the waveform to stabilize
        time.sleep(2)

        # Capture waveform
        sc.run()
        time.sleep((load_settings['low_time'] + load_settings['high_time']) / 1e6 * 5)  # Capture multiple cycles
        # Analyze results
        vout_max = float(sc.get_max_measurement(Index=1))
        vout_min = float(sc.get_min_measurement(Index=2))
        rise_time = float(sc.get_val_measurement(Index=3))
        fall_time = float(sc.get_val_measurement(Index=4))

        # Convert overshoot and undershoot to mV
        overshoot_mv = vout_max * 1000
        undershoot_mv = vout_min * 1000

        # Convert rise and fall times to µs
        rise_time_us = rise_time * 1e6
        fall_time_us = fall_time * 1e6

        print("Capturing full ...")
        # Save full waveform image
        sc.set_screen_text(text=f"I low: {load_settings['i_low']}A, I high: {load_settings['i_high']}A", text_no=1, xpos="10", ypos="20")
        sc.set_screen_text(text=f"Low time: {load_settings['low_time']}us, High time: {load_settings['high_time']}us", text_no=2, xpos="10", ypos="25")
        full_cycle_image = os.path.join(result_folder, f"{test_name}_full_cycle.png")
        sc.saveimage(full_cycle_image)
        
        # Capture rising edge
        print("Capturing rising edge...")
        sc._write(f'HOR:SCALE 10e-6')  # Zoom in 10x
        sc.Trigger_Edge(Level=trigger_level, Slope='RISE', Coupling='DC', ChannelIndex=int(iout_ch_num[-1]))
        sc.run()
        time.sleep(1)
        sc.set_screen_text(text="Rising Edge", text_no=3, xpos="10", ypos="30")
        rising_edge_image = os.path.join(result_folder, f"{test_name}_rising_edge.png")
        sc.saveimage(rising_edge_image)
        
        # Capture falling edge
        print("Capturing falling edge...")
        sc._write(f'HOR:SCALE 10e-6')  # Zoom in 10x
        sc.Trigger_Edge(Level=trigger_level, Slope='FALL', Coupling='DC', ChannelIndex=int(iout_ch_num[-1]))
        sc.run()
        time.sleep(1)
        sc.set_screen_text(text="Falling Edge", text_no=3, xpos="10", ypos="30")
        falling_edge_image = os.path.join(result_folder, f"{test_name}_falling_edge.png")
        sc.saveimage(falling_edge_image)
        
        # Reset scope settings
        sc._write(f'HOR:SCALE {scope_us_div}e-6')
        sc._write('HOR:POS 50')
        sc.set_screen_text(text="", text_no=3)

     

 
        vout_ch.ProbeSetup(Coupling='DC', Bandwidth='20E+06', Vrange=10, Impedance=1e+6, Position=-3.5)
        # Allow some time for the change to take effect
        time.sleep(0.5)

        # Set up measurements

        sc.set_measurement(MeasureIndex=3, MeasureType='MEAN', State='ON', Source=int(vout_ch_num[-1]))

        
        # Capture waveform
        sc.run()
        time.sleep(1)  # Allow time for measurement to stabilize

        # Get measurements

        vout_mean = float(sc.get_mean_measurement(Index=3))

        
        # Calculate overshoot and undershoot
        overshoot = vout_max*1000
        undershoot = -1*vout_mean - vout_min*1000

        # Calculate percentages
        overshoot_percentage = (vout_max / vout_mean) * 100
        undershoot_percentage = (-vout_min / vout_mean) * 100

        # Determine pass/fail status
        overshoot_pass = overshoot_percentage <= pass_fail_criteria['overshoot']
        undershoot_pass = undershoot_percentage <= pass_fail_criteria['undershoot']

        # Print results for debugging
        print(f"Vout mean: {vout_mean:.3f} V")
        print(f"Overshoot: {overshoot:.3f} mV ({overshoot_percentage:.2f}%)  {'pass' if overshoot_pass else 'fail'}")
        print(f"Undershoot: {undershoot:.3f} mV ({undershoot_percentage:.2f}%)  {'pass' if undershoot_pass else 'fail'}")

        results = {
            'vout_mean': f"{vout_mean:.3f} V" ,
            'overshoot': f"{overshoot:.3f} V ({overshoot_percentage:.2f}%)",
            'undershoot': f"{undershoot:.3f} V ({undershoot_percentage:.2f}%)",
            'rise_time': f"{rise_time*1e6:.2f} µs",
            'fall_time': f"{fall_time*1e6:.2f} µs",
            'overshoot_pass': overshoot_pass,
            'undershoot_pass': undershoot_pass,
        }


        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Transient Test Results"

        # Add test information
        ws['A1'] = "Transient Test Results"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A3'] = f"Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"IC: {selected_ic}"

        # Add settings information
        ws['A6'] = "Test Settings"
        ws['A6'].font = Font(size=14, bold=True)
        ws['A7'] = f"Input Voltage: {power_supply_settings['vin']} V"
        ws['A8'] = f"Input Current Limit: {power_supply_settings['iin']} A"
        ws['A9'] = f"Low Load Current: {load_settings['i_low']} A"
        ws['A10'] = f"High Load Current: {load_settings['i_high']} A"
        ws['A11'] = f"Low Time: {load_settings['low_time']} µs"
        ws['A12'] = f"High Time: {load_settings['high_time']} µs"
        ws['A13'] = f"Overshoot Limit: {pass_fail_criteria['overshoot']}%"
        ws['A14'] = f"Undershoot Limit: {pass_fail_criteria['undershoot']}%"
        # Add results
        ws['A16'] = "Test Results"
        ws['A16'].font = Font(size=14, bold=True)
        ws['A17'] = f"Rise Time: {results['rise_time']}"
        ws['A18'] = f"Fall Time: {results['fall_time']}"
        ws['A19'] = f"Overshoot: {results['overshoot']}"
        ws['A20'] = f"Undershoot: {results['undershoot']}"
        
        # Define fill colors
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

        # Add pass/fail indicators with color
        ws['B19'] = 'PASS' if results['overshoot_pass'] else 'FAIL'
        ws['B19'].fill = green_fill if results['overshoot_pass'] else red_fill
        ws['B19'].font = Font(color='FFFFFF', bold=True)
        ws['B19'].alignment = Alignment(horizontal='center')

        ws['B20'] = 'PASS' if results['undershoot_pass'] else 'FAIL'
        ws['B20'].fill = green_fill if results['undershoot_pass'] else red_fill
        ws['B20'].font = Font(color='FFFFFF', bold=True)
        ws['B20'].alignment = Alignment(horizontal='center')

        # Adjust column widths for better readability
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15


        # Add images
        img_row = 30
        for img_file, title in [(full_cycle_image, "Full Cycle"), (rising_edge_image, "Rising Edge"), (falling_edge_image, "Falling Edge")]:
            img = Image(img_file)
            img.width = 1200
            img.height = 600
            ws.add_image(img, f'A{img_row}')
            ws[f'A{img_row-3}'] = title
            ws[f'A{img_row-3}'].font = Font(size=12, bold=True)
            img_row += 40

        # Save Excel file
        wb.save(excel_file)

        print("Transient test completed")
        print(f"Results saved to {excel_file}")

        
        # Delete image files
        for img_file in [full_cycle_image, rising_edge_image, falling_edge_image]:
            if os.path.exists(img_file):
                os.remove(img_file)

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        import traceback
        print(traceback.format_exc())
        set_stop_flag()
    finally:
        if 'electronic_load' in locals():
            electronic_load.output('OFF')
        if 'power_supply_channel' in locals():
            power_supply_channel.Enable(False)
        if 'sc' in locals():
            sc.set_screen_text(text="", text_no=1)
            sc.set_screen_text(text="", text_no=2)
        print("Test ended, instruments turned off")

    return results



