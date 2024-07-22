import os
import time
from datetime import datetime
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from config import *
import pyvisa
from pverifyDrivers import _scope, powsup, daq, load

def switching_node(selected_ic, power_supply_settings, scope_settings, scope_persistence, scope_us_div, load_settings, protection, instrument_manager):
    print(f"Debug: switching_node function started for {selected_ic}")
    results = {}
    print("Switching Node test started")

    # Create result folder
    result_folder = os.path.join("results", "switching_node")
    os.makedirs(result_folder, exist_ok=True)

    # Generate unique filename for this test
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    test_name = f"Switching_Node_Test_{selected_ic}_{timestamp}"
    excel_file = os.path.join(result_folder, f"{test_name}.xlsx")

    try:
        instruments = instrument_manager.initialize_instruments()
        
        electronic_load = instruments['load']
        power_supply = instruments['supply']
        sc = instruments['o_scope']
        
        power_supply_channel = power_supply.GetChannel(power_supply_settings['vin_channel'])

        # Find channel numbers for SW, Vout, and Iout
        sw_ch_num = next(ch for ch, signal in scope_settings.items() if signal == 'SW')
        vout_ch_num = next(ch for ch, signal in scope_settings.items() if signal == 'output V')
        iout_ch_num = next(ch for ch, signal in scope_settings.items() if signal == 'output I')

        # Setup oscilloscope channels
        sw_ch = sc.GetChannel(int(sw_ch_num[-1]))
        vout_ch = sc.GetChannel(int(vout_ch_num[-1]))
        iout_ch = sc.GetChannel(int(iout_ch_num[-1]))

        # Enable channels
        sw_ch.Enable(True)
        vout_ch.Enable(True)
        iout_ch.Enable(True)

        # Set up channels with provided settings
        sw_ch.ProbeSetup(Coupling='DC', Bandwidth='20E+06', Vrange=10, Impedance=50, Position=0)
        vout_ch.ProbeSetup(Coupling='DC', Bandwidth='20E+06', Vrange=10, Impedance=1e+6, Position=-3)
        iout_ch.ProbeSetup(Coupling='DC', Bandwidth='20E+06', Vrange=10, Impedance=1e+6, Position=3)

        # Set channel labels
        sc.set_channel_label(ChannelIndex=int(sw_ch_num[-1]), Label="SW", Xpos="5", Ypos="80")
        sc.set_channel_label(ChannelIndex=int(vout_ch_num[-1]), Label="Vout", Xpos="5", Ypos="50")
        sc.set_channel_label(ChannelIndex=int(iout_ch_num[-1]), Label="Iout", Xpos="5", Ypos="20")

        # Set horizontal scale
        sc._write(f'HOR:SCALE {scope_us_div}e-6')

        sc.set_acquisition_mode('HIRes')
        sc.set_record_length(10000000)  # 10M points for higher resolution

        # Function to capture and save waveform
        def capture_and_save_waveform(filename, text):
            sc.run()
            time.sleep(1)
            sc.set_screen_text(text=text, text_no=1, xpos="10", ypos="20")
            image_path = os.path.join(result_folder, filename)
            sc.saveimage(image_path)
            return image_path

        for vin in power_supply_settings['vin']:
            # Set up power supply for current Vin
            power_supply_channel.Configure_VoltageLevel(Level=vin, CurrentLimit=power_supply_settings['iin'])
            power_supply_channel.Enable(True)
            time.sleep(1)  # Allow time for voltage to stabilize

            # Set trigger on the SW channel
            sc.Trigger_Edge(Level=vin/2, Slope='RISE', Coupling='DC', ChannelIndex=int(sw_ch_num[-1]))

            results[f'Vin_{vin}V'] = {}

            # 1. Fsw vs. Iout
            results[f'Vin_{vin}V']['fsw_vs_iout'] = []
            for iout in load_settings['load_values']:
                electronic_load.set_current(value=iout)
                time.sleep(load_settings['load_delay'])
                sc.set_measurement(MeasureIndex=1, MeasureType='FREQUENCY', State='ON', Source=int(sw_ch_num[-1]))
                fsw = float(sc.get_val_measurement(Index=1))
                results[f'Vin_{vin}V']['fsw_vs_iout'].append((iout, fsw))

            # 2. Ton vs. Iout
            results[f'Vin_{vin}V']['ton_vs_iout'] = []
            for iout in load_settings['load_values']:
                electronic_load.set_current(value=iout)
                time.sleep(load_settings['load_delay'])
                sc.set_measurement(MeasureIndex=2, MeasureType='PWIDTH', State='ON', Source=int(sw_ch_num[-1]))
                ton = float(sc.get_val_measurement(Index=2))
                results[f'Vin_{vin}V']['ton_vs_iout'].append((iout, ton))

            # 3. SW node peak (Rising and Falling)
            sc.set_measurement(MeasureIndex=3, MeasureType='MAXIMUM', State='ON', Source=int(sw_ch_num[-1]))
            sc.set_measurement(MeasureIndex=4, MeasureType='MINIMUM', State='ON', Source=int(sw_ch_num[-1]))
            results[f'Vin_{vin}V']['sw_peak_rising'] = float(sc.get_val_measurement(Index=3))
            results[f'Vin_{vin}V']['sw_peak_falling'] = float(sc.get_val_measurement(Index=4))

            # 4. Dead time (Rising and Falling)
            dead_time_image = capture_and_save_waveform(f"{test_name}_Vin_{vin}V_dead_time.png", f"Dead Time Measurement (Vin={vin}V)")
            results[f'Vin_{vin}V']['dead_time_image'] = dead_time_image

            # 5. Jitter
            sc.set_measurement(MeasureIndex=5, MeasureType='PJITTER', State='ON', Source=int(sw_ch_num[-1]))
            results[f'Vin_{vin}V']['jitter'] = float(sc.get_val_measurement(Index=5))

            # Additional measurements
            # 1. Switch node ringing, peak voltages
            ringing_image = capture_and_save_waveform(f"{test_name}_Vin_{vin}V_ringing.png", f"Switch Node Ringing (Vin={vin}V)")
            results[f'Vin_{vin}V']['ringing_image'] = ringing_image

            # 2. Rise time and fall times
            sc.set_measurement(MeasureIndex=6, MeasureType='RISETIME', State='ON', Source=int(sw_ch_num[-1]))
            sc.set_measurement(MeasureIndex=7, MeasureType='FALLTIME', State='ON', Source=int(sw_ch_num[-1]))
            results[f'Vin_{vin}V']['rise_time'] = float(sc.get_val_measurement(Index=6))
            results[f'Vin_{vin}V']['fall_time'] = float(sc.get_val_measurement(Index=7))

            # 3. dV/dT distortion
            dvdt_image = capture_and_save_waveform(f"{test_name}_Vin_{vin}V_dvdt.png", f"dV/dT Distortion (Vin={vin}V)")
            results[f'Vin_{vin}V']['dvdt_image'] = dvdt_image

        # Create Excel file and save results
        wb = Workbook()
        ws = wb.active
        ws.title = "Switching Node Test Results"

        # Add test information and results to Excel
        ws['A1'] = "Switching Node Test Results"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A3'] = f"Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A4'] = f"IC: {selected_ic}"

        row = 6
        for vin, vin_results in results.items():
            ws[f'A{row}'] = vin
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            for key, value in vin_results.items():
                ws[f'A{row}'] = key
                if isinstance(value, list):
                    ws[f'B{row}'] = str(value)
                elif isinstance(value, str) and value.endswith('.png'):
                    img = Image(value)
                    img.width = 300
                    img.height = 200
                    ws.add_image(img, f'B{row}')
                    row += 15  # Add space for the image
                else:
                    ws[f'B{row}'] = str(value)
                row += 1
            row += 1  # Add a blank row between Vin results

        # Save Excel file
        wb.save(excel_file)

        print("Switching Node test completed")
        print(f"Results saved to {excel_file}")

    except Exception as e:
        print(f"An error occurred: {str(e)}")
        import traceback
        print(traceback.format_exc())
        set_stop_flag()
    finally:
        if 'electronic_load' in locals():
            electronic_load.output('OFF')
        if 'power_supply_channel' in locals():
            power_supply_channel.Enable(False)
        if 'sc' in locals():
            sc.set_screen_text(text="", text_no=1)
        print("Test ended, instruments turned off")

    return results